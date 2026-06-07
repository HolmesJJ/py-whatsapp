import os
import sys
import time
import argparse
import pywhatkit

from datetime import datetime
from openpyxl import load_workbook


DEFAULT_WAIT = 40
MASTER_FILE = 'Master.xlsx'
BACKUP_DIR = 'backups'
ARGS_FILE = 'DO NOT TOUCH.txt'


def test():
    pywhatkit.sendwhatmsg_instantly('+65', 'Hi, how are you?', DEFAULT_WAIT, True)
    pywhatkit.sendwhatmsg_instantly('+65', 'Hi, how are you?', DEFAULT_WAIT, True)


def get_folder_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def make_backup(wb_write, master_path):
    folder = os.path.join(os.path.dirname(master_path), BACKUP_DIR)
    os.makedirs(folder, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = os.path.splitext(MASTER_FILE)[0]
    backup_name = f"{base_name}_{ts}.xlsx"
    backup_path = os.path.join(folder, backup_name)
    wb_write.save(backup_path)
    print(f'✓ Backup saved as {os.path.join(BACKUP_DIR, backup_name)}')
    all_backups = [
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.startswith(base_name) and f.endswith('.xlsx')
    ]
    if len(all_backups) > 5:
        all_backups.sort(key=os.path.getmtime, reverse=True)
        to_delete = all_backups[5:]
        for old_file in to_delete:
            try:
                os.remove(old_file)
            except Exception as e:
                print(f'❌ Failed to delete {os.path.basename(old_file)}: {e}')


def run(wait=DEFAULT_WAIT):
    print('--- Start ---')
    folder_path = get_folder_path()
    master_path = os.path.join(folder_path, MASTER_FILE)
    if not os.path.exists(master_path):
        print(f'❌ {MASTER_FILE} not found in {folder_path}.')
        print('--- End ---')
        return
    try:
        print(f'--- Processing Excel {MASTER_FILE} ... ---')
        wb_read = load_workbook(master_path, data_only=True)
        wb_write = load_workbook(master_path, data_only=False)
        ws_r, ws_w = wb_read.worksheets[0], wb_write.worksheets[0]
        header = list(next(ws_r.iter_rows(min_row=1, max_row=1, values_only=True)))
        try:
            idx_id = header.index('ID') + 1
            idx_mobile = header.index('Mobile Number') + 1
            idx_name = header.index('Official Name') + 1
            idx_message = header.index('Full Text Message') + 1
            idx_sent = header.index('Sent') + 1
        except ValueError as e2:
            print(f'❌ Missing columns: {e2}')
            wb_read.close()
            wb_write.close()
            print('--- End ---')
            return
        for row_r, row_w in zip(ws_r.iter_rows(min_row=2), ws_w.iter_rows(min_row=2)):
            id_cell = row_r[idx_id - 1]
            mobile_cell = row_r[idx_mobile - 1]
            name_cell = row_r[idx_name - 1]
            message_cell = row_r[idx_message - 1]
            sent_cell = row_w[idx_sent - 1]
            if not mobile_cell.value:
                print(f'❌ Row {id_cell.value} missing mobile value')
                break
            if not name_cell.value:
                print(f'❌ Row {id_cell.value} missing name value')
                break
            if not message_cell.value:
                print(f'❌ Row {id_cell.value} missing message value')
                break
            if sent_cell.value == 1:
                continue
            mobile = str(mobile_cell.value).strip()
            message = str(message_cell.value).strip()
            name = str(name_cell.value).strip()
            full_number = mobile if mobile.startswith('+') else f'+{mobile}'
            print()
            print(f'→ {name} | {full_number} | wait={wait}s | message={message[:30]}...')
            try:
                pywhatkit.sendwhatmsg_instantly(full_number, message, wait, True)
                sent_cell.value = 1
                print(f'! Updating Excel file: {MASTER_FILE}. Please DO NOT close or terminate the program...')
                wb_write.save(master_path)
                make_backup(wb_write, master_path)
                print(f'✓ Excel file {MASTER_FILE} has been updated. '
                      f'You may now safely close or terminate the program...')
                print('→ Waiting 5 seconds...', end='', flush=True)
                for i in range(5, 0, -1):
                    print(f' {i}...', end='', flush=True)
                    time.sleep(1)
                print()
                print()
            except Exception as e3:
                print(f'❌ Send failed: {e3}')
        wb_read.close()
        wb_write.close()
    except Exception as e1:
        print(f'❌ Read {MASTER_FILE} error: {e1}')
    print('--- End ---')


def load_args_file():
    """Read key=value pairs from args.txt next to the executable/script."""
    folder = get_folder_path()
    args_path = os.path.join(folder, ARGS_FILE)
    config = {}
    if os.path.exists(args_path):
        print(f'--- Reading {ARGS_FILE} ---')
        with open(args_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                if '=' in line:
                    key, value = line.split('=', 1)
                    config[key.strip()] = value.strip()
    else:
        print(f'⚠ {ARGS_FILE} not found in {folder}, using defaults.')
    return config


def parse_args():
    file_config = load_args_file()
    parser = argparse.ArgumentParser(description='Bulk-send WhatsApp messages based on Excel files.')
    parser.add_argument('--wait', type=int,
                        default=int(file_config.get('wait', DEFAULT_WAIT)),
                        help=f'Wait (seconds) before sending. The default is {DEFAULT_WAIT} seconds.')
    return parser.parse_args()


if __name__ == '__main__':
    # pyinstaller --onefile --console --hidden-import=pywhatkit --hidden-import=openpyxl --name "DO NOT TOUCH" main.py
    args = parse_args()
    print('--- Args ---')
    print(f'  wait = {args.wait}s')
    print('-------------')
    print()
    run(args.wait)
